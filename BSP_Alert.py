import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from openpyxl.styles.borders import Border, Side
from pathlib import Path
import json
import spacy
# import en_core_web_sm
# import spacy_streamlit
# from spacy_streamlit import load_model
from openpyxl.drawing.image import Image
from Tools import title_clean, bg_image


def sheet_formating(df, sendout_date):

    # category cell formats
    category_color_fill = PatternFill(start_color='06A2E5', end_color='06A2E5', fill_type='solid')
    category_font = Font(bold=True, color='FFFFFF', name='Arial', size=12)

    # header cell formats
    header_font = Font(bold=True, name='Arial', size=10)

    # general formats
    thin_border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))
    
    general_font = Font(bold=False, name='Arial', size=10)

    sendout_font = Font(bold=True, name='Arial', size=20)
    
    BSP_FILE = Path(__file__).parent/f'templates/BSP_Temp/bsp_template.xlsx'

    # st.dataframe(df)

    df_cat1 = df.groupby('CATEGORY').get_group('TODAYS HEADLINENEWS')

    # get rows whose ONLINE LINK and PRINT LINK has values
    df_1 = df_cat1.loc[df_cat1['ONLINE LINK'] != 'N/A']
    df_1 = df_1.loc[df_1['PRINT LINK'] != 'N/A']

    # get rows whose print rows
    df_2 = df_cat1.loc[df_cat1['ONLINE LINK'] == 'N/A']

    # get online rows
    df_3 = df_cat1.loc[df_cat1['PRINT LINK'] == 'N/A']

    # combine dataframes
    df_cat1 = pd.concat([df_1, df_2, df_3], sort=False)
    df_cat1.reset_index(drop=True, inplace=True)

    df_cat2 = df.groupby('CATEGORY').get_group('TODAYS BUSINESS HEADLINENEWS')

    # get rows whose ONLINE LINK and PRINT LINK has values
    df_1 = df_cat2.loc[df_cat2['ONLINE LINK'] != 'N/A']
    df_1 = df_1.loc[df_1['PRINT LINK'] != 'N/A']

    # get rows whose print rows
    df_2 = df_cat2.loc[df_cat2['ONLINE LINK'] == 'N/A']

    # get online rows
    df_3 = df_cat2.loc[df_cat2['PRINT LINK'] == 'N/A']

    # combine dataframes
    df_cat2 = pd.concat([df_1, df_2, df_3], sort=False)
    df_cat2.reset_index(drop=True, inplace=True)

    df_cat3 = df.groupby('CATEGORY').get_group('BSP NEWS')

    # get rows whose ONLINE LINK and PRINT LINK has values
    df_1 = df_cat3.loc[df_cat3['ONLINE LINK'] != 'N/A']
    df_1 = df_1.loc[df_1['PRINT LINK'] != 'N/A']

    # get rows whose print rows
    df_2 = df_cat3.loc[df_cat3['ONLINE LINK'] == 'N/A']

    # get online rows
    df_3 = df_cat3.loc[df_cat3['PRINT LINK'] == 'N/A']

    # combine dataframes
    df_cat3 = pd.concat([df_1, df_2, df_3], sort=False)
    df_cat3.reset_index(drop=True, inplace=True)

    dfs = []
    dfs.append(df_cat1)
    dfs.append(df_cat2)
    dfs.append(df_cat3)

    wb = openpyxl.Workbook()
    ws = wb.active

    sendout_date_cell = ws.cell(row=1, column=1)
    sendout_date_cell.value = sendout_date
    sendout_date_cell.font = sendout_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    ws.cell(row=5, column=1).value = 'Kindly click on the following links to view your respective news. For best result in viewing the links your default browser should be set to Google Chrome, Mozilla Firefox, Internet Explorer version 10 or higher.'

    # insert bsp logo
    image_path = Path(__file__).parent/f'templates/BSP_Temp/bsp_logo.jpg'
    bsp_logo = Image(image_path)
    ws.add_image(bsp_logo, 'A3')

    l, w = df_cat1.shape
    
    cats = ['TODAYS HEADLINE NEWS', 'TODAYS BUSINESS HEADLINENEWS', 'BSP NEWS']
    c = 0
    s_row = 8
    for _df in dfs:
        
        l, w = _df.shape
        
        # create category header
        category_cell = ws.cell(row=s_row, column=1)
        category_cell.value = cats[c]
        category_cell.fill = category_color_fill
        category_cell.border = thin_border
        category_cell.font = category_font
        category_cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=s_row, start_column=1, end_row=s_row, end_column=5)

        # create header
        date_header = ws.cell(row=s_row+1, column=1)
        date_header.value = 'DATE'
        date_header.font = header_font
        date_header.alignment = Alignment(horizontal='center')
        date_header.border = thin_border

        source_header = ws.cell(row=s_row+1, column=2)
        source_header.value = 'SOURCE'
        source_header.font = header_font
        source_header.alignment = Alignment(horizontal='center')
        source_header.border = thin_border

        title_header = ws.cell(row=s_row+1, column=3)
        title_header.value = 'TITLE'
        title_header.font = header_font
        title_header.alignment = Alignment(horizontal='center')
        title_header.border = thin_border

        online_header = ws.cell(row=s_row+1, column=4)
        online_header.value = 'ONLINE'
        online_header.font = header_font
        online_header.alignment = Alignment(horizontal='center')
        online_header.border = thin_border

        print_header = ws.cell(row=s_row+1, column=5)
        print_header.value = 'PRINT'
        print_header.font = header_font
        print_header.alignment = Alignment(horizontal='center')
        print_header.border = thin_border
        
        for i in _df.index:

            s_date = _df.at[i, 'DATE']
            s_source = _df.at[i, 'SOURCE']
            s_title = _df.at[i, 'TITLE']
            s_online = _df.at[i, 'ONLINE LINK']
            s_print = _df.at[i, 'PRINT LINK']    

            # populate and format cells
            date_cell = ws.cell(row=s_row+2+i, column=1)
            date_cell.value = s_date
            date_cell.number_format = 'MMM DD, YYYY'
            date_cell.alignment = Alignment(horizontal='center')
            date_cell.font = general_font
            date_cell.border = thin_border
            
            source_cell = ws.cell(row=s_row+2+i, column=2)
            source_cell.value = s_source
            source_cell.alignment = Alignment(horizontal='center')
            source_cell.font = general_font
            source_cell.border = thin_border

            title_cell = ws.cell(row=s_row+2+i, column=3)
            title_cell.value = s_title
            title_cell.font = general_font
            title_cell.border = thin_border

            online_cell = ws.cell(row=s_row+2+i, column=4)
            if s_online == 'N/A':
                online_cell.value = 'N/A'
                online_cell.font = general_font
            else:
                online_cell.value = 'Online Link'
                online_cell.hyperlink = s_online
                # online_cell.style = 'Hyperlink'
                online_cell.font = Font(bold=True, name='Arial', size=10, color='0000FF', underline='single')
            
            online_cell.alignment = Alignment(horizontal='center')
            online_cell.border = thin_border

            print_cell = ws.cell(row=s_row+2+i, column=5)
            if s_print == 'N/A':    
                print_cell.value = 'N/A'
                print_cell.font = general_font
            else:
                print_cell.value = 'Print Link'
                print_cell.hyperlink = s_print
                # print_cell.style = 'Hyperlink'
                print_cell.font = Font(bold=True, name='Arial', size=10, color='0000FF', underline='single')

            print_cell.alignment = Alignment(horizontal='center')
            print_cell.border = thin_border


        s_row = s_row + l + 2
        c += 1

    wb.save(BSP_FILE)
    wb.close()

    return BSP_FILE



def json_publications():

    f = open('json_files/bsp_publications.json')
    bsp_pub = json.load(f)

    return bsp_pub





def similar_title(nlp, a, b):
    a1 = nlp(a)
    b1 = nlp(b)
    return a1.similarity(b1)




def dataframe_create(uploaded_file):

    REPORT_FILE = Path(__file__).parent/f'templates/BSP_Temp/trial.xlsx'
    wb = openpyxl.Workbook(REPORT_FILE)
    wb.save(REPORT_FILE)
    wb.close()

    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active

    sendout_date = ws.cell(row=1, column=1).value

    s_row = 1
    for row in ws.iter_rows():
        try:
            ws.unmerge_cells(start_row=s_row, start_column=1, end_row=s_row, end_column=7)
        except:
            pass
        s_row += 1

    

    for row in ws.iter_rows(max_col=7):

        if row[0].value == 'TODAYS HEADLINENEWS':
            cat = 'TODAYS HEADLINENEWS'
        elif row[0].value == 'TODAYS BUSINESS HEADLINENEWS':
            cat = 'TODAYS BUSINESS HEADLINENEWS'
        elif row[0].value == 'BSPNEWS':
            cat = 'BSP NEWS'
        
        if row[2].hyperlink != None:
            row[4].value = row[6].value
            row[5].value = cat
            row[6].value = row[2].hyperlink.target
            row[2].hyperlink = None

    for row in ws.iter_rows(max_col=7):

        if row[0].value in ['TODAYS HEADLINENEWS',
                            'TODAYS BUSINESS HEADLINENEWS',
                            'BSPNEWS']:
            row[0].value = ''

    ws.delete_rows(1,5)
    wb.save(REPORT_FILE)
    wb.close()

    df = pd.read_excel(REPORT_FILE) 

    return df, REPORT_FILE, sendout_date


def bsp_main():

    bg_image()

    with st.container(border=True):
        
        st.header('Bangko Sentral ng Pilipinas')

    st.file_uploader('Input Raw File', key='bsp_raw')

    if st.session_state['bsp_raw'] != None:
        button_process = st.button('Process File')

        if button_process:

            nlp = spacy.load('en_core_web_sm')
            # nlp = load_model('en_core_web_sm')
            # nlp = en_core_web_sm.load()

            df, REPORT_FILE, sendout_date = dataframe_create(st.session_state['bsp_raw'])
            
            df.columns = ['DATE', 'SOURCE', 'TITLE', 'AUTHOR', 'TYPE', 'CATEGORY', 'LINK']

            df['PRINT LINK'] = 'N/A'
            df['ONLINE LINK'] = 'N/A'
            df['DELETE'] = ''

            df_cat1 = df.groupby('CATEGORY').get_group('TODAYS HEADLINENEWS')

            for i in df_cat1.index:
                _title = df_cat1.at[i, 'TITLE']
                
                if _title[:10].lower() == 'headlines:':
                    df_cat1.at[i, 'TITLE'] = _title[11:]
                    _title = df_cat1.at[i, 'TITLE']
                
                if '|' in _title:
                    end_text = _title.split('|')[-1]
                    df_cat1.at[i, 'TITLE'] = _title.replace(end_text, '')

            df_cat2 = df.groupby('CATEGORY').get_group('TODAYS BUSINESS HEADLINENEWS')

            for i in df_cat2.index:
                _title = df_cat2.at[i, 'TITLE']
                
                if _title[:10].lower() == 'headlines:':
                    df_cat2.at[i, 'TITLE'] = _title[11:]
                    _title = df_cat2.at[i, 'TITLE']
                
                if '|' in _title:
                    end_text = _title.split('|')[-1]
                    df_cat2.at[i, 'TITLE'] = _title.replace(end_text, '')

            df_cat3 = df.groupby('CATEGORY').get_group('BSP NEWS')

            for i in df_cat3.index:
                _title = df_cat3.at[i, 'TITLE']
                
                if _title[:10].lower() == 'headlines:':
                    df_cat3.at[i, 'TITLE'] = _title[11:]
                    _title = df_cat3.at[i, 'TITLE']
                
                if '|' in _title:
                    end_text = _title.split('|')[-1]
                    df_cat3.at[i, 'TITLE'] = _title.replace(end_text, '')

            dfs = []
            dfs.append(df_cat1)
            dfs.append(df_cat2)
            dfs.append(df_cat3)
            
            # load print/online counterparts
            bsp = json_publications()

            broadsheet_list = []
            for k, v in bsp.items():
                broadsheet_list.append(k)

            new_dfs = []
            for _df in dfs:
                for j in _df.index:
                    main_title = _df.at[j, 'TITLE']
                    main_source = _df.at[j, 'SOURCE']
                    main_link = _df.at[j, 'LINK']
                    main_type = _df.at[j, 'TYPE']
                    
                    if main_type in ['Online News', 'Blogs']:
                        _df.at[j, 'ONLINE LINK'] = main_link
                        continue
                    elif main_type in ['Magazine']:
                        _df.at[j, 'PRINT LINK'] = main_link
                        _df.at[j, 'DELETE'] = 'DONE'
                        continue
                    elif main_source not in broadsheet_list:
                        _df.at[j, 'PRINT LINK'] = main_link
                        _df.at[j, 'DELETE'] = 'DONE'
                        continue
                    else:
                        _df.at[j, 'PRINT LINK'] = main_link

                        for k in _df.index:
                            sub_title = _df.at[k, 'TITLE']
                            sub_source = _df.at[k, 'SOURCE']
                            sub_link = _df.at[k, 'LINK']
                            sub_type = _df.at[k, 'TYPE']
                            sub_delete = _df.at[k, 'DELETE']

                            if sub_type != 'Online News':
                                continue
                            else:
                                if sub_source not in bsp[main_source]:
                                    continue
                                else:
                                    if sub_delete == 'FOR DELETION':
                                        continue
                                    else:
                                        try:
                                            sub_title = title_clean(sub_title, sub_source)
                                        except:
                                            pass

                                        similarity_ratio = similar_title(nlp, main_title.lower(), sub_title.lower())
                                        if similarity_ratio < 0.8:
                                            continue
                                        else:
                                            _df.at[j, 'ONLINE LINK'] = sub_link
                                            _df.at[k, 'DELETE'] = 'FOR DELETION'
                                            break

                        _df.at[j, 'DELETE'] = 'DONE'                    

                # drop rows for deletion
                _df = _df[_df.DELETE != 'FOR DELETION']

                # drop other columns
                _df = _df.drop(["AUTHOR", "LINK", "DELETE"], axis='columns')

                # st.dataframe(_df)
                new_dfs.append(_df)


            for df in new_dfs:
                df['TYPE'] = pd.Categorical(df['TYPE'], ['Broadsheet', 'Tabloid', 'Provincial', 'Magazine', 'Online News', 'Blogs'])
                df.sort_values('TYPE')

            # convert to excel
            df_merged = pd.concat([new_dfs[0], new_dfs[1], new_dfs[2]], sort=False)
            df_merged = df_merged[['DATE', 'SOURCE', 'TITLE', 'ONLINE LINK', 'PRINT LINK', 'CATEGORY']]

            df_merged.to_excel(REPORT_FILE, index=False, startrow=8)
            
            BSP_FILE = sheet_formating(df_merged, sendout_date)

            result_file = open(BSP_FILE, 'rb')
            st.success(f':red[NOTE:] Downloaded file will go to the :red[Downloads Folder]')
            st.download_button(label='📥 Download Excel File', data= result_file, file_name= f'bsp_template.xlsx')

         
