import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from pathlib import Path
import json
import spacy

def sheet_formating(df):

    BSP_FILE = Path(__file__).parent/f'BSP_Temp/bsp_template.xlsx'

    # st.dataframe(df)

    df_cat1 = df.groupby('CATEGORY').get_group('TODAYS HEADLINENEWS')
    df_cat1.reset_index(drop=True, inplace=True)
    df_cat2 = df.groupby('CATEGORY').get_group('TODAYS BUSINESS HEADLINENEWS')
    df_cat2.reset_index(drop=True, inplace=True)
    df_cat3 = df.groupby('CATEGORY').get_group('BSP NEWS')
    df_cat3.reset_index(drop=True, inplace=True)

    dfs = []
    dfs.append(df_cat1)
    dfs.append(df_cat2)
    dfs.append(df_cat3)

    st.dataframe(df_cat1)
    # st.dataframe(df_cat2)
    # st.dataframe(df_cat3)

    wb = openpyxl.Workbook()
    ws = wb.active

    l, w = df_cat1.shape
    
    cats = ['TODAYS HEADLINE NEWS', 'TODAYS BUSINESS HEADLINENEWS', 'BSP NEWS']
    c = 0
    s_row = 8
    for _df in dfs:
        # st.write(s_row)
        l, w = _df.shape
        # st.write(l)
        ws.cell(row=s_row, column=1).value = cats[c]
        ws.cell(row=s_row+1, column=1).value = 'DATE'
        ws.cell(row=s_row+1, column=2).value = 'SOURCE'
        ws.cell(row=s_row+1, column=3).value = 'TITLE'
        ws.cell(row=s_row+1, column=4).value = 'ONLINE'
        ws.cell(row=s_row+1, column=5).value = 'PRINT'
        
        for i in _df.index:
            ws.cell(row=s_row+2+i, column=1).value = _df.at[i, 'DATE']
            ws.cell(row=s_row+2+i, column=2).value = _df.at[i, 'SOURCE']
            ws.cell(row=s_row+2+i, column=3).value = _df.at[i, 'TITLE']
            ws.cell(row=s_row+2+i, column=4).value = _df.at[i, 'ONLINE LINK']
            ws.cell(row=s_row+2+i, column=5).value = _df.at[i, 'PRINT LINK']
            st.write(i)
        s_row = s_row + l + 3
        c += 1

    wb.save(BSP_FILE)
    wb.close()

    return BSP_FILE



def json_publications():

    f = open('json_files/bsp_publications.json')
    bsp_pub = json.load(f)

    return bsp_pub


def similar_title(a, b):
    a1 = nlp(a)
    b1 = nlp(b)
    return a1.similarity(b1)




def dataframe_create(uploaded_file):

    REPORT_FILE = Path(__file__).parent/f'BSP_Temp/trial.xlsx'
    wb = openpyxl.Workbook(REPORT_FILE)
    wb.save(REPORT_FILE)
    wb.close()

    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active


    s_row = 1
    for row in ws.iter_rows():
        try:
            ws.unmerge_cells(start_row=s_row, start_column=1, end_row=s_row, end_column=7)
        except:
            pass
        s_row += 1

    

    for row in ws.iter_rows(max_col=7):

        if row[0].value == 'TODAYS HEADLINENEWS':
            cat = 'TODAYS HEADLINENEWS'
        elif row[0].value == 'TODAYS BUSINESS HEADLINENEWS':
            cat = 'TODAYS BUSINESS HEADLINENEWS'
        elif row[0].value == 'BSPNEWS':
            cat = 'BSP NEWS'
        
        if row[2].hyperlink != None:
            row[4].value = row[6].value
            row[5].value = cat
            row[6].value = row[2].hyperlink.target
            row[2].hyperlink = None

    for row in ws.iter_rows(max_col=7):

        if row[0].value in ['TODAYS HEADLINENEWS',
                            'TODAYS BUSINESS HEADLINENEWS',
                            'BSPNEWS']:
            row[0].value = ''

    ws.delete_rows(1,7)
    wb.save(REPORT_FILE)
    wb.close()

    df = pd.read_excel(REPORT_FILE) 

    return df, REPORT_FILE




with st.container(border=True):
    st.header('TRIAL')

st.file_uploader('Input Raw File', key='bsp_raw')

if st.session_state['bsp_raw'] != None:
    button_process = st.button('Process File')

    if button_process:

        nlp = spacy.load('en_core_web_sm')

        df, REPORT_FILE = dataframe_create(st.session_state['bsp_raw'])

        df.columns = ['DATE', 'SOURCE', 'TITLE', 'AUTHOR', 'TYPE', 'CATEGORY', 'LINK']

        df['PRINT LINK'] = 'N/A'
        df['ONLINE LINK'] = 'N/A'
        df['DELETE'] = ''

        df_cat1 = df.groupby('CATEGORY').get_group('TODAYS HEADLINENEWS')
        df_cat2 = df.groupby('CATEGORY').get_group('TODAYS BUSINESS HEADLINENEWS')
        df_cat3 = df.groupby('CATEGORY').get_group('BSP NEWS')

        dfs = []
        dfs.append(df_cat1)
        dfs.append(df_cat2)
        dfs.append(df_cat3)
        
        # load print/online counterparts
        bsp = json_publications()

        broadsheet_list = []
        for k, v in bsp.items():
            broadsheet_list.append(k)

        new_dfs = []
        for _df in dfs:
            for j in _df.index:
                main_title = _df.at[j, 'TITLE']
                main_source = _df.at[j, 'SOURCE']
                main_link = _df.at[j, 'LINK']
                main_type = _df.at[j, 'TYPE']
                
                if main_type in ['Online News', 'Blogs']:
                    _df.at[j, 'ONLINE LINK'] = main_link
                    continue
                elif main_type in ['Tabloid', 'Magazine', 'Provincial']:
                    _df.at[j, 'PRINT LINK'] = main_link
                    _df.at[j, 'DELETE'] = 'DONE'
                    continue
                elif main_source not in broadsheet_list:
                    _df.at[j, 'PRINT LINK'] = main_link
                    _df.at[j, 'DELETE'] = 'DONE'
                    continue
                else:
                    _df.at[j, 'PRINT LINK'] = main_link

                    for k in _df.index:
                        sub_title = _df.at[k, 'TITLE']
                        sub_source = _df.at[k, 'SOURCE']
                        sub_link = _df.at[k, 'LINK']
                        sub_type = _df.at[k, 'TYPE']
                        sub_delete = _df.at[k, 'DELETE']

                        if sub_type != 'Online News':
                            continue
                        else:
                            if sub_source not in bsp[main_source]:
                                continue
                            else:
                                if sub_delete == 'FOR DELETION':
                                    continue
                                else:
                                    # try:
                                        # st.write(sub_title)
                                        # sub_title = sub_title.replace('BusinessWorld Online', '')
                                        # st.write(sub_title)
                                    # except:
                                    #     pass

                                    similarity_ratio = similar_title(main_title.lower(), sub_title.lower())
                                    if similarity_ratio < 0.8:
                                        continue
                                    else:
                                        _df.at[j, 'ONLINE LINK'] = sub_link
                                        _df.at[k, 'DELETE'] = 'FOR DELETION'
                                        break

                    _df.at[j, 'DELETE'] = 'DONE'                    

            # drop rows for deletion
            _df = _df[_df.DELETE != 'FOR DELETION']

            # drop other columns
            _df = _df.drop(["AUTHOR", "LINK", "DELETE"], axis='columns')

            # st.dataframe(_df)
            new_dfs.append(_df)


        for df in new_dfs:
            df['TYPE'] = pd.Categorical(df['TYPE'], ['Broadsheet', 'Tabloid', 'Provincial', 'Magazine', 'Online News', 'Blogs'])
            df.sort_values('TYPE')

        # convert to excel
        df_merged = pd.concat([new_dfs[0], new_dfs[1], new_dfs[2]], sort=False)
        df_merged = df_merged[['DATE', 'SOURCE', 'TITLE', 'ONLINE LINK', 'PRINT LINK', 'CATEGORY']]

        df_merged.to_excel(REPORT_FILE, index=False, startrow=8)
        
        BSP_FILE = sheet_formating(df_merged)

        result_file = open(BSP_FILE, 'rb')
        st.success(f':red[NOTE:] Downloaded file will go to the :red[Downloads Folder]')
        st.download_button(label='📥 Download Excel File', data= result_file, file_name= f'bsp_template.xlsx')

         
