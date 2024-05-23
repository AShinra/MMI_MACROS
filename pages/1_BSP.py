import streamlit as st
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from openpyxl.styles.borders import Border, Side
from pathlib import Path




with st.container(border=True):
    st.header('Bangko Sentral ng Pilipinas Macro')

st.file_uploader('Input Raw File', key='bsp_raw')

if st.session_state['bsp_raw'] != None:
    button_process = st.button('Process File')

    if button_process:

        REPORT_FILE = Path(__file__).parent/f'BSP_Temp/bsp.xlsx'
        wb = openpyxl.Workbook(REPORT_FILE)
        wb.save(REPORT_FILE)
        wb.close()

        wb = openpyxl.load_workbook(st.session_state['bsp_raw'])
        ws = wb.active

        s_row = 8
        for row in ws.iter_rows(min_row=8, max_col=7):

            row[0].alignment = Alignment(horizontal='center')
            color_fill = PatternFill(start_color='06A2E5', end_color='06A2E5', fill_type='solid')
            thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

            if row[0].value=='TODAYS HEADLINENEWS':
                ws.unmerge_cells(start_row=s_row, start_column=1, end_row=s_row, end_column=7)
                ws.merge_cells(start_row=s_row, start_column=1, end_row=s_row, end_column=6)
                active_cell = row[0]
                active_cell.value = 'TODAYS HEADLINE NEWS'
                active_cell.fill = color_fill
                active_cell.alignment = Alignment(horizontal='center')
            
            if row[0].value=='TODAYS BUSINESS HEADLINENEWS':
                ws.unmerge_cells(start_row=s_row, start_column=1, end_row=s_row, end_column=7)
                ws.merge_cells(start_row=s_row, start_column=1, end_row=s_row, end_column=6)
                active_cell = row[0]
                active_cell.value = 'TODAYS BUSINESS HEADLINE NEWS'
                active_cell.fill = color_fill
                active_cell.alignment = Alignment(horizontal='center')
            
            if row[0].value=='BSPNEWS':
                ws.unmerge_cells(start_row=s_row, start_column=1, end_row=s_row, end_column=7)
                ws.merge_cells(start_row=s_row, start_column=1, end_row=s_row, end_column=6)
                active_cell = row[0]
                active_cell.value = 'BSP NEWS'
                active_cell.fill = color_fill
                active_cell.alignment = Alignment(horizontal='center')

            if row[0].value=='DATE':
                row[1].value='SOURCE'
                row[2].value='TITLE'
                row[3].value='AUTHOR'
                row[4].value='PRINT'
                row[5].value='ONLINE'

            if row[2].hyperlink != None:

                if row[6].value=='Online News':
                    row[4].value='N/A'
                    row[4].alignment = Alignment(horizontal='center')
                    row[5].value=row[2].hyperlink.target
                    row[5].hyperlink=row[5].value
                    row[5].value='Online Link'
                    row[5].style = 'Hyperlink'
                    row[5].alignment = Alignment(horizontal='center')
                    row[5].border = thin_border
                else:
                    row[5].value='N/A'
                    row[5].alignment = Alignment(horizontal='center')
                    row[4].value=row[2].hyperlink.target
                    row[4].hyperlink=row[4].value
                    row[4].value='Print Link'
                    row[4].style = 'Hyperlink'
                    row[4].alignment = Alignment(horizontal='center')
                    row[4].border = thin_border

            if row[2].value != None and row[2].value != 'TITLE':
                row[2].hyperlink = None
                active_cell = row[2]
                ft = Font(color='000000')
                active_cell.font = ft
            
            ws.row_dimensions[s_row].height = 15
            s_row += 1            

        ws.delete_cols(7,1)
        wb.save(REPORT_FILE)
        wb.close()

        result_file = open(REPORT_FILE, 'rb')
        st.success(f':red[NOTE:] Downloaded file will go to the :red[Downloads Folder]')
        st.download_button(label='📥 Download Excel File', data= result_file, file_name= f'bsp.xlsx')


         

